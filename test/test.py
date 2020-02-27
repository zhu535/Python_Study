# # -*- coding: utf-8 -*-
# # @Time : 2020/2/25 11:42
# # @Author : zhu
# # @FileName: test.py
# # @Software: PyCharm

import os

import openpyxl
import pandas as pd
import xlwt,xlrd
from xlutils.copy import copy

# 将文件读取出来放一个列表里面

pwd = 'excel'  # 获取文件目录

# 新建列表，存放文件名
file_list = []

# 新建列表存放每个文件数据(依次读取多个相同结构的Excel文件并创建DataFrame)
dfs = []

for root, dirs, files in os.walk(pwd):  # 第一个为起始路径，第二个为起始路径下的文件夹，第三个是起始路径下的文件。
    for file in files:
        file_path = os.path.join(root, file)
        print("正在读取%s的内容" %file_path)
        file_list.append(file_path)  # 使用os.path.join(dirpath, name)得到全路径
        df = pd.read_excel(file_path,sheet_name='sheet1',header=None,skiprows=8)  # 将excel转换成DataFrame
        dfs.append(df)

# 将多个DataFrame合并为一个
df = pd.concat(dfs,sort=False)

# 写入excel文件，不包含索引数据
df.to_excel('data.xlsx', header=False, index=False)

read_title = xlrd.open_workbook(os.path.join(os.getcwd(), 'title.xls'),formatting_info=True)
sheet1 = read_title.sheet_by_index(0)

result = copy(read_title)
table = result.get_sheet(0)

data = openpyxl.load_workbook(os.path.join(os.getcwd(), 'data.xlsx'))
data_sheet1 = data['Sheet1']
mc = data_sheet1.max_column
mr = data_sheet1.max_row

for i in range(1, mr+1):
    for j in range(1, mc+1):
        value = data_sheet1.cell(i, j).value
        table.write(i+7, j-1, value)

result.save(os.path.join(os.getcwd(), '雄金.XLS'))

